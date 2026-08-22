import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment

FONT = "Arial"
BLUE = Font(name=FONT, color="0000FF")
BLACK = Font(name=FONT, color="000000")
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
BOLD_WHITE = Font(name=FONT, bold=True, color="FFFFFF")
TITLE = Font(name=FONT, bold=True, size=14)
SECTION = Font(name=FONT, bold=True, size=11)
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CUR = '$#,##0;($#,##0);"-"'
PCT = '0.0%'

wb = openpyxl.Workbook()

def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# ============================================================
# Sheet 1: Assumptions
# ============================================================
ws = wb.active
ws.title = "Assumptions"
set_col_widths(ws, [46, 14, 44])

ws["A1"] = "Meridian Health Network — Cost Model Assumptions"
ws["A1"].font = TITLE
ws["A2"] = "All figures are illustrative, order-of-magnitude Azure retail (Pay-As-You-Go, East US) list-price approximations for 3–5 year TCO comparison purposes only — not a vendor quote, EA/MCA negotiated rate, or budget-approval figure. Blue cells are editable inputs."
ws["A2"].font = Font(name=FONT, italic=True, size=9, color="666666")
ws.merge_cells("A2:C2")
ws["A2"].alignment = Alignment(wrap_text=True)
ws.row_dimensions[2].height = 30

row = 4
ws.cell(row=row, column=1, value="General").font = SECTION
row += 1
general_rows = {}
def add_input(ws, row, label, value, note, numfmt=None):
    ws.cell(row=row, column=1, value=label).font = BLACK
    c = ws.cell(row=row, column=2, value=value)
    c.font = BLUE
    if numfmt:
        c.number_format = numfmt
    ws.cell(row=row, column=3, value=note).font = Font(name=FONT, size=9, color="666666")
    return row

r = row
r = add_input(ws, r, "Workforce headcount", 4300, "current-state.md §1"); general_rows["workforce"] = r; r += 1
r = add_input(ws, r, "Active patients (last 24 months)", 410000, "current-state.md §1"); general_rows["patients"] = r; r += 1
r = add_input(ws, r, "Sites", 46, "current-state.md §1"); general_rows["sites"] = r; r += 1
r = add_input(ws, r, "Hours per year", 8760, "24 x 365"); general_rows["hours"] = r; r += 1
r = add_input(ws, r, "Reserved Capacity / Savings Plan discount (Year 2+, ADR-038)", 0.25, "Illustrative 1-yr Reserved Instance / Savings Plan discount range on reservable compute/DB lines only", PCT); general_rows["ri_discount"] = r; r += 1

r += 1
ws.cell(row=r, column=1, value="On-Prem Baseline ('Do-Nothing') Inputs").font = SECTION
r += 1
r = add_input(ws, r, "Capital refresh cost — SAN + 6 hosts + networking (one-time)", 660000, "current-state.md §3 (91% SAN util, 6.5yr-avg host fleet); illustrative", CUR); general_rows["capex"] = r; r += 1
r = add_input(ws, r, "Capital refresh amortization period (years)", 5, "Straight-line, for TCO comparison only"); general_rows["capex_years"] = r; r += 1
r = add_input(ws, r, "Backup / tape-courier / after-hours MSP NOC, annual", 150000, "current-state.md §3/§1; illustrative", CUR); general_rows["backup_msp"] = r; r += 1
r = add_input(ws, r, "MPLS WAN to 46 sites, annual", 248400, "current-state.md §4; illustrative ~$450/site/month", CUR); general_rows["mpls"] = r; r += 1
r = add_input(ws, r, "Cyber-insurance baseline annual premium", 180000, "problem-statement.md §4; illustrative", CUR); general_rows["ins_base"] = r; r += 1
r = add_input(ws, r, "Cyber-insurance penalty multiplier if MFA/backup/DR conditions unmet", 3, 'problem-statement.md §4: "up to 3x current cost"'); general_rows["ins_mult"] = r; r += 1

r += 1
ws.cell(row=r, column=1, value="Migration Program Inputs (migration-roadmap.md, ADR-035/036/037)").font = SECTION
r += 1
r = add_input(ws, r, "Program duration, Wave 0 through Wave 10 (months)", 13, "migration-roadmap.md diagram"); general_rows["duration"] = r; r += 1
r = add_input(ws, r, "Azure run-rate ramp — average % of steady-state during Year 1", 0.40, "Illustrative wave-weighted average; most sites cut over in the back half of the program", PCT); general_rows["ramp"] = r; r += 1
r = add_input(ws, r, "Legacy on-prem opex continuation during Year 1 (% of full-year cost)", 1.0, "ADR-036: legacy farm stays fully operational as rollback fallback until Wave 10", PCT); general_rows["legacy_pct"] = r; r += 1

ASM = general_rows  # dict of row numbers on Assumptions sheet

# ============================================================
# Sheet 2: Azure Run-Rate (Steady State)
# ============================================================
ws2 = wb.create_sheet("AzureRunRate")
set_col_widths(ws2, [34, 30, 8, 12, 12, 14, 10, 42])
ws2["A1"] = "Azure Target-State Run-Rate — Steady State (Post Wave 10, Both Regions Where Applicable)"
ws2["A1"].font = TITLE
ws2.merge_cells("A1:H1")

hdr = ["Service / Line Item", "Configuration", "Qty", "Unit Rate ($)", "Billing Basis", "Annual Cost", "Reservable?", "Source / Note"]
for i, h in enumerate(hdr, start=1):
    ws2.cell(row=3, column=i, value=h)
style_header_row(ws2, 3, len(hdr))

rows_data = [
    ("CareLink PM Citrix VDA compute", "Standard_D8s_v5, Windows Datacenter (ADR-005)", 20, 1.00, "Hourly", "Y",
     "ADR-005 sizing (day-1 catalog 45, autoscale floor ~7); 20 = illustrative average running fleet across business/off-hours"),
    ("SQL Managed Instance — primary (East US)", "Business Critical, 8 vCore, Zone Redundant Gen5 (ADR-006)", 8, 0.78, "Hourly", "Y",
     "ADR-006 sizing; rate = illustrative Business Critical Zone-Redundant vCore-hr incl. storage"),
    ("SQL Managed Instance — DR (West US)", "Matching Business Critical instance, auto-failover group (ADR-006)", 8, 0.78, "Hourly", "Y",
     "ADR-004/006: warm standby must match primary, not a smaller stand-in"),
    ("App Service Premium v3 (Portal)", "P1v3 Linux, autoscale 3-10 (ADR-010)", 5, 0.146, "Hourly", "Y",
     "ADR-010 sizing; 5 = illustrative average across the 3-10 autoscale range"),
    ("Azure Functions Premium (LinkEngine)", "EP1 x4 apps, shared plan, 1 Always Ready each (ADR-008)", 4, 0.173, "Hourly", "Y",
     "ADR-008 sizing; excludes minor consumption-based burst execution"),
    ("Service Bus Premium — primary", "1 Messaging Unit (ADR-011)", 1, 677, "Monthly", "N",
     "ADR-011 sizing"),
    ("Service Bus Premium — Geo-DR secondary", "1 Messaging Unit, West US namespace (ADR-011)", 1, 677, "Monthly", "N",
     "ADR-011: Geo-DR secondary namespace billed at full rate (topology-only replication)"),
    ("API Management Premium — primary", "1 scale unit (Section 1 service mapping)", 1, 2795, "Monthly", "N",
     "Premium tier required for VNet integration + Strangler Fig ingress"),
    ("API Management Premium — DR scale unit", "1 additional scale unit, West US", 1, 2795, "Monthly", "N",
     "Multi-region APIM deployment for DR"),
    ("Azure Front Door Premium", "Base + WAF + data transfer (ADR-010)", 1, 667, "Monthly", "N",
     "Illustrative base + usage blend"),
    ("Azure Site Recovery", "Golden-image / management-tier protected instances (ADR-004/§8)", 5, 25, "Monthly", "N",
     "VDA session hosts are ephemeral/non-persistent (ADR-005) and rebuilt from image on failover, not individually ASR-protected"),
    ("Blob/Backup storage (GRS + immutable, both regions)", "~200 TB blended hot/cool/archive tiers", 200000, 0.02, "Monthly-per-GB", "N",
     "Includes ADR-037 Wave-0 immutable retention-lock backup target"),
    ("Microsoft Sentinel (SIEM)", "~50 GB/day ingestion, combined Sentinel + Log Analytics", 50, 2.46, "Daily-per-GB", "N",
     "Illustrative commit-tier blended rate"),
    ("Microsoft Entra ID P2 (workforce Conditional Access/MFA)", "Per-user/month, all workforce", 4300, 9.00, "Monthly-per-user", "N",
     "ADR-009/ADR-037; LARGEST single line — verify against any existing M365/EMS E5 licensing before budgeting (not resolved here)"),
    ("Microsoft Entra External ID (patient CIAM)", "Monthly Active Users beyond free tier", 73000, 0.01, "Monthly-per-MAU", "N",
     "ADR-009; ~30% of active patients assumed to engage portal monthly, less a 50,000-MAU free tier"),
    ("Retained hybrid connectivity (post-decommission)", "Modest ExpressRoute circuit for any remaining HQ-only, out-of-scope systems", 1, 1600, "Monthly", "N",
     "Temporary migration-window circuit (ADR-035) is separate — see MigrationYr1"),
    ("Hub networking — Azure Firewall + Bastion + DNS (both regions)", "Standard tier, primary + DR hub", 1, 2083, "Monthly", "N",
     "azure-implementation.md hub-and-spoke; illustrative blended both-region rate"),
    ("Miscellaneous / contingency", "Key Vault transactions, DNS zones, minor services", 1, 833, "Monthly", "N",
     "Buffer line, not itemized"),
]

start = 4
for i, (svc, cfg, qty, rate, basis, reservable, note) in enumerate(rows_data):
    rr = start + i
    ws2.cell(row=rr, column=1, value=svc).font = BLACK
    ws2.cell(row=rr, column=2, value=cfg).font = BLACK
    qc = ws2.cell(row=rr, column=3, value=qty); qc.font = BLUE
    rc = ws2.cell(row=rr, column=4, value=rate); rc.font = BLUE; rc.number_format = '$#,##0.00'
    ws2.cell(row=rr, column=5, value=basis).font = BLACK
    fcell = ws2.cell(row=rr, column=6)
    if basis == "Hourly":
        fcell.value = f"=C{rr}*D{rr}*Assumptions!$B${ASM['hours']}"
    elif basis == "Monthly":
        fcell.value = f"=C{rr}*D{rr}*12"
    elif basis == "Monthly-per-GB":
        fcell.value = f"=C{rr}*D{rr}*12"
    elif basis == "Daily-per-GB":
        fcell.value = f"=C{rr}*D{rr}*365"
    elif basis == "Monthly-per-user":
        fcell.value = f"=C{rr}*D{rr}*12"
    elif basis == "Monthly-per-MAU":
        fcell.value = f"=C{rr}*D{rr}*12"
    fcell.font = BLACK
    fcell.number_format = CUR
    ws2.cell(row=rr, column=7, value=reservable).font = BLACK
    ws2.cell(row=rr, column=7).alignment = Alignment(horizontal="center")
    notecell = ws2.cell(row=rr, column=8, value=note)
    notecell.font = Font(name=FONT, size=9, color="666666")
    notecell.alignment = Alignment(wrap_text=True)
    for c in range(1, 9):
        ws2.cell(row=rr, column=c).border = BORDER

last_data_row = start + len(rows_data) - 1
total_row = last_data_row + 2
ws2.cell(row=total_row, column=1, value="Total — Azure Steady-State Annual Run-Rate").font = BOLD
ws2.cell(row=total_row, column=6, value=f"=SUM(F{start}:F{last_data_row})")
ws2.cell(row=total_row, column=6).font = BOLD
ws2.cell(row=total_row, column=6).number_format = CUR
for c in range(1, 9):
    ws2.cell(row=total_row, column=c).fill = TOTAL_FILL
    ws2.cell(row=total_row, column=c).border = BORDER

reservable_row = total_row + 1
ws2.cell(row=reservable_row, column=1, value="  of which: Reservable (RI/Savings-Plan-eligible) compute & DB lines").font = Font(name=FONT, italic=True, size=9)
ws2.cell(row=reservable_row, column=6, value=f'=SUMIF(G{start}:G{last_data_row},"Y",F{start}:F{last_data_row})')
ws2.cell(row=reservable_row, column=6).number_format = CUR
ws2.cell(row=reservable_row, column=6).font = Font(name=FONT, italic=True, size=9)

RR = {"total_row": total_row, "reservable_row": reservable_row}

# ============================================================
# Sheet 3: On-Prem Baseline (Status Quo)
# ============================================================
ws3 = wb.create_sheet("OnPremBaseline")
set_col_widths(ws3, [46, 16, 44])
ws3["A1"] = "On-Prem Baseline — 'Do Nothing' / Stay-and-Refresh Scenario (Annual, Flat Across Years)"
ws3["A1"].font = TITLE
ws3.merge_cells("A1:C1")

hdr3 = ["Cost Component", "Annual Cost", "Source / Note"]
for i, h in enumerate(hdr3, start=1):
    ws3.cell(row=3, column=i, value=h)
style_header_row(ws3, 3, len(hdr3))

r = 4
items = [
    ("Capital refresh, amortized (SAN + hosts + networking)", f"=Assumptions!$B${ASM['capex']}/Assumptions!$B${ASM['capex_years']}",
     "requirements.md: capital refresh 'due in the current on-prem model regardless'"),
    ("Backup / tape-courier / after-hours MSP NOC", f"=Assumptions!$B${ASM['backup_msp']}", "current-state.md §3"),
    ("MPLS WAN to 46 sites", f"=Assumptions!$B${ASM['mpls']}", "current-state.md §4"),
    ("Cyber-insurance premium (penalty applies — current architecture cannot meet MFA/immutable-backup/tested-DR conditions)",
     f"=Assumptions!$B${ASM['ins_base']}*Assumptions!$B${ASM['ins_mult']}", "problem-statement.md §4"),
]
onprem_start = r
for label, formula, note in items:
    ws3.cell(row=r, column=1, value=label).font = BLACK
    ws3.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    fc = ws3.cell(row=r, column=2, value=formula); fc.font = GREEN; fc.number_format = CUR
    nc = ws3.cell(row=r, column=3, value=note); nc.font = Font(name=FONT, size=9, color="666666"); nc.alignment = Alignment(wrap_text=True)
    for c in range(1, 4):
        ws3.cell(row=r, column=c).border = BORDER
    r += 1
onprem_end = r - 1
onprem_total_row = r + 1
ws3.cell(row=onprem_total_row, column=1, value="Total — On-Prem Baseline, Annual").font = BOLD
ws3.cell(row=onprem_total_row, column=2, value=f"=SUM(B{onprem_start}:B{onprem_end})")
ws3.cell(row=onprem_total_row, column=2).font = BOLD
ws3.cell(row=onprem_total_row, column=2).number_format = CUR
for c in range(1, 4):
    ws3.cell(row=onprem_total_row, column=c).fill = TOTAL_FILL
    ws3.cell(row=onprem_total_row, column=c).border = BORDER

note_row = onprem_total_row + 2
ws3.cell(row=note_row, column=1, value="Note: staffing (16-person team) is unchanged between scenarios and excluded here as a wash — see cost-and-risk-analysis.md §1.").font = Font(name=FONT, italic=True, size=9, color="666666")
ws3.merge_cells(f"A{note_row}:C{note_row}")
ws3.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)

OP = {"total_row": onprem_total_row, "backup_msp_ref": f"OnPremBaseline!B{onprem_start+1}", "mpls_ref": f"OnPremBaseline!B{onprem_start+2}"}

# ============================================================
# Sheet 4: Migration Transition (Year 1)
# ============================================================
ws4 = wb.create_sheet("MigrationYr1")
set_col_widths(ws4, [46, 16, 44])
ws4["A1"] = "Migration & Transition Costs — Year 1 Only (Wave 0 through Wave 10)"
ws4["A1"].font = TITLE
ws4.merge_cells("A1:C1")

hdr4 = ["Cost Component", "Cost", "Source / Note"]
for i, h in enumerate(hdr4, start=1):
    ws4.cell(row=3, column=i, value=h)
style_header_row(ws4, 3, len(hdr4))

r = 4
ws4.cell(row=r, column=1, value="One-time transition costs").font = SECTION
r += 1
onetime_start = r
onetime = [
    ("Landing zone build & IaC development (Wave 0)", 180000, "Bicep modules, network, identity foundation — azure-implementation.md §14 deferred item, built here"),
    ("Migration execution labor, all waves (~$42K/wave average x 10 waves)", 420000, "migration-roadmap.md wave count; per-wave cutover engineering, testing, smoke-test gates"),
    ("Temporary hybrid ExpressRoute/VPN circuit (13 months)", 32500, "ADR-035's named networking dependency between Wave 1 and the last compute wave"),
    ("Citrix Cloud incremental licensing, dual-farm overlap period", 60000, "Both legacy and Azure Citrix fleets run simultaneously per ADR-036"),
    ("Clinician / front-desk training and change management", 150000, "migration-roadmap.md §6, explicitly deferred workstream, budgeted here"),
]
for label, val, note in onetime:
    ws4.cell(row=r, column=1, value=label).font = BLACK
    ws4.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    vc = ws4.cell(row=r, column=2, value=val); vc.font = BLUE; vc.number_format = CUR
    nc = ws4.cell(row=r, column=3, value=note); nc.font = Font(name=FONT, size=9, color="666666"); nc.alignment = Alignment(wrap_text=True)
    for c in range(1, 4):
        ws4.cell(row=r, column=c).border = BORDER
    r += 1
onetime_end = r - 1
onetime_subtotal_row = r
ws4.cell(row=r, column=1, value="Subtotal — one-time").font = BOLD
ws4.cell(row=r, column=2, value=f"=SUM(B{onetime_start}:B{onetime_end})").font = BOLD
ws4.cell(row=r, column=2).number_format = CUR
for c in range(1, 4):
    ws4.cell(row=r, column=c).fill = SUBHEADER_FILL
    ws4.cell(row=r, column=c).border = BORDER
r += 2

ws4.cell(row=r, column=1, value="Ramped / continuing costs during Year 1").font = SECTION
r += 1
ramped_start = r
ramped = [
    ("Azure run-rate, ramped (wave-weighted average of steady state)",
     f"=AzureRunRate!F{RR['total_row']}*Assumptions!$B${ASM['ramp']}",
     "Most sites cut over in the back half of the program (migration-roadmap.md wave timeline)"),
    ("Legacy on-prem opex continuation (backup/MSP + MPLS, full year — ADR-036)",
     f"=({OP['backup_msp_ref']}+{OP['mpls_ref']})*Assumptions!$B${ASM['legacy_pct']}",
     "ADR-036: legacy farm stays fully operational as rollback fallback until Wave 10"),
    ("Capital refresh — AVOIDED (migrating replaces the need for it entirely, not just defers it)", 0,
     "Real business-case point: this is $0, not deferred spend — see cost-and-risk-analysis.md"),
    ("Cyber-insurance premium, Year 1 (baseline rate preserved)",
     f"=Assumptions!$B${ASM['ins_base']}",
     "ADR-037: MFA/encryption/immutable-backup quick-wins land in Wave 0, weeks into the program — no penalty"),
]
for label, val, note in ramped:
    ws4.cell(row=r, column=1, value=label).font = BLACK
    ws4.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    vc = ws4.cell(row=r, column=2, value=val)
    vc.number_format = CUR
    vc.font = GREEN if isinstance(val, str) else BLUE
    nc = ws4.cell(row=r, column=3, value=note); nc.font = Font(name=FONT, size=9, color="666666"); nc.alignment = Alignment(wrap_text=True)
    for c in range(1, 4):
        ws4.cell(row=r, column=c).border = BORDER
    r += 1
ramped_end = r - 1
ramped_subtotal_row = r
ws4.cell(row=r, column=1, value="Subtotal — ramped/continuing").font = BOLD
ws4.cell(row=r, column=2, value=f"=SUM(B{ramped_start}:B{ramped_end})").font = BOLD
ws4.cell(row=r, column=2).number_format = CUR
for c in range(1, 4):
    ws4.cell(row=r, column=c).fill = SUBHEADER_FILL
    ws4.cell(row=r, column=c).border = BORDER
r += 2

yr1_total_row = r
ws4.cell(row=r, column=1, value="Total — Azure-Track Cost, Year 1").font = BOLD
ws4.cell(row=r, column=2, value=f"=B{onetime_subtotal_row}+B{ramped_subtotal_row}").font = BOLD
ws4.cell(row=r, column=2).number_format = CUR
for c in range(1, 4):
    ws4.cell(row=r, column=c).fill = TOTAL_FILL
    ws4.cell(row=r, column=c).border = BORDER

MIG = {"yr1_total_row": yr1_total_row}

# ============================================================
# Sheet 5: TCO Summary
# ============================================================
ws5 = wb.create_sheet("TCOSummary")
set_col_widths(ws5, [30, 13, 13, 13, 13, 13, 13, 13])
ws5["A1"] = "5-Year (+2 Illustrative) TCO Comparison — Do-Nothing Baseline vs. Azure Migration"
ws5["A1"].font = TITLE
ws5.merge_cells("A1:H1")
ws5["A2"] = "Years 6-7 are shown only to illustrate the breakeven trend beyond the requested 3-5 year window — see cost-and-risk-analysis.md §3 for the honest read on what this comparison does and doesn't show."
ws5["A2"].font = Font(name=FONT, italic=True, size=9, color="666666")
ws5.merge_cells("A2:H2")
ws5.row_dimensions[2].height = 26
ws5["A2"].alignment = Alignment(wrap_text=True)

years = ["Year 1", "Year 2", "Year 3", "Year 4", "Year 5", "Year 6", "Year 7"]
hdr5 = ["Scenario"] + years
for i, h in enumerate(hdr5, start=1):
    ws5.cell(row=4, column=i, value=h)
style_header_row(ws5, 4, len(hdr5))

# Row 5: Do-Nothing annual
r = 5
ws5.cell(row=r, column=1, value="Do-Nothing Baseline (annual)").font = BLACK
for ci in range(2, 9):
    cell = ws5.cell(row=r, column=ci, value=f"=OnPremBaseline!$B${OP['total_row']}")
    cell.font = GREEN
    cell.number_format = CUR
    cell.border = BORDER
donothing_row = r

r += 1
ws5.cell(row=r, column=1, value="Azure Migration (annual)").font = BLACK
azure_row = r
# Year1 -> MigrationYr1 total; Year2+ -> RunRate total minus RI discount on reservable subset
ws5.cell(row=r, column=2, value=f"=MigrationYr1!$B${MIG['yr1_total_row']}").font = GREEN
ws5.cell(row=r, column=2).number_format = CUR
for ci in range(3, 9):
    col = openpyxl.utils.get_column_letter(ci)
    formula = (f"=AzureRunRate!$F${RR['total_row']}"
               f"-AzureRunRate!$F${RR['reservable_row']}*Assumptions!$B${ASM['ri_discount']}")
    cell = ws5.cell(row=r, column=ci, value=formula)
    cell.font = GREEN
    cell.number_format = CUR
for ci in range(2, 9):
    ws5.cell(row=r, column=ci).border = BORDER

r += 2
ws5.cell(row=r, column=1, value="Cumulative Do-Nothing").font = BOLD
cum_dn_row = r
for ci in range(2, 9):
    col = openpyxl.utils.get_column_letter(ci)
    prev_col = openpyxl.utils.get_column_letter(ci - 1)
    if ci == 2:
        formula = f"=B{donothing_row}"
    else:
        formula = f"={prev_col}{r}+{col}{donothing_row}"
    cell = ws5.cell(row=r, column=ci, value=formula)
    cell.font = BOLD
    cell.number_format = CUR
    cell.border = BORDER

r += 1
ws5.cell(row=r, column=1, value="Cumulative Azure Migration").font = BOLD
cum_az_row = r
for ci in range(2, 9):
    col = openpyxl.utils.get_column_letter(ci)
    prev_col = openpyxl.utils.get_column_letter(ci - 1)
    if ci == 2:
        formula = f"=B{azure_row}"
    else:
        formula = f"={prev_col}{r}+{col}{azure_row}"
    cell = ws5.cell(row=r, column=ci, value=formula)
    cell.font = BOLD
    cell.number_format = CUR
    cell.border = BORDER

r += 1
ws5.cell(row=r, column=1, value="Cumulative Delta (Azure minus Do-Nothing)").font = Font(name=FONT, italic=True)
delta_row = r
for ci in range(2, 9):
    col = openpyxl.utils.get_column_letter(ci)
    cell = ws5.cell(row=r, column=ci, value=f"={col}{cum_az_row}-{col}{cum_dn_row}")
    cell.font = Font(name=FONT, italic=True)
    cell.number_format = CUR
    cell.border = BORDER

for c in range(1, 9):
    ws5.cell(row=cum_dn_row, column=c).fill = SUBHEADER_FILL
    ws5.cell(row=cum_az_row, column=c).fill = SUBHEADER_FILL
    ws5.cell(row=delta_row, column=c).fill = TOTAL_FILL

# Chart: cumulative comparison
chart = BarChart()
chart.type = "col"
chart.title = "Cumulative Cost — Do-Nothing vs. Azure Migration"
chart.y_axis.title = "Cumulative Cost ($)"
chart.x_axis.title = "Year"
data = Reference(ws5, min_col=2, max_col=8, min_row=cum_dn_row, max_row=cum_az_row)
cats = Reference(ws5, min_col=2, max_col=8, min_row=4, max_row=4)
chart.add_data(data, titles_from_data=False, from_rows=True)
chart.set_categories(cats)
chart.series[0].tx = openpyxl.chart.series.SeriesLabel(v="Do-Nothing Baseline")
chart.series[1].tx = openpyxl.chart.series.SeriesLabel(v="Azure Migration")
chart.height = 9
chart.width = 22
ws5.add_chart(chart, f"A{delta_row + 3}")

wb.save("/home/claude/work/Cloud-Architecture/case-study-03-healthcare-platform/finance/TCO-Analysis.xlsx")
print("saved")
